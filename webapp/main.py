"""
库存整理工具的网页后端。

单页应用：上传一个新的库存快照xlsx，后端自动识别格式（260624_style/erp_style）、
提取快照日期、登记进 raw_data/snapshot_registry.json，然后跑一遍
scripts/clean_latest.py 的整理流程，把结果xlsx存到 output/，页面上直接展示分析
摘要（品牌数量、库存状态分布、临期预警、低库存预警），并提供完整结果下载链接。

每次成功整理的结果会按快照日期归档到 output/history/{date}.html 和
output/history/{date}.xlsx，左侧栏列出所有归档过的日期，点击可以查看/下载
历史上某一期的整理结果，不会因为后面又传了新快照就丢掉旧记录。

启动方式（本地测试）：
    cd Inventory/webapp
    uvicorn main:app --host 0.0.0.0 --port 8000
"""
import html
import io
import json
import shutil
import sys
from datetime import datetime
from pathlib import Path

import pandas as pd
from fastapi import FastAPI, File, Form, HTTPException, UploadFile
from fastapi.responses import FileResponse, HTMLResponse, StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

PROJECT_ROOT = Path(__file__).resolve().parent.parent
SCRIPTS_DIR = PROJECT_ROOT / "scripts"
RAW_DIR = PROJECT_ROOT / "raw_data"
OUTPUT_DIR = PROJECT_ROOT / "output"
HISTORY_DIR = OUTPUT_DIR / "history"
SALES_DIR = RAW_DIR / "sales"

sys.path.insert(0, str(SCRIPTS_DIR))

import clean_latest  # noqa: E402
from categorize import STATUS_COLORS  # noqa: E402
from load_sales import load_monthly_sales, get_sales_months  # noqa: E402

app = FastAPI(title="库存整理工具")

PAGE_TEMPLATE = """<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>库存整理工具</title>
<style>
  * {{ box-sizing: border-box; }}
  body {{ font-family: -apple-system, "Microsoft YaHei", sans-serif; margin: 0; color: #1f2937; background: #fafafa; }}
  .layout {{ display: flex; min-height: 100vh; }}
  .sidebar {{ width: 220px; flex-shrink: 0; background: #1f2937; color: #d1d5db; padding: 20px 0; }}
  .sidebar h3 {{ font-size: 13px; color: #9ca3af; padding: 0 16px; margin: 0 0 8px; text-transform: uppercase; }}
  .sidebar a {{ display: block; padding: 8px 16px; color: #d1d5db; text-decoration: none; font-size: 14px; }}
  .sidebar a:hover {{ background: #374151; }}
  .sidebar a.active {{ background: #2563eb; color: #fff; font-weight: 600; }}
  .sidebar .empty {{ padding: 0 16px; font-size: 13px; color: #6b7280; }}
  .main {{ flex: 1; max-width: 980px; margin: 32px auto; padding: 0 16px; }}
  h1 {{ font-size: 22px; }}
  h2 {{ font-size: 16px; margin: 28px 0 10px; }}
  .card {{ border: 1px solid #e5e7eb; border-radius: 8px; padding: 24px; margin-top: 16px; background: #fff; }}
  input[type=file] {{ margin: 12px 0; }}
  button {{ background: #1f2937; color: #fff; border: none; padding: 10px 20px; border-radius: 6px; cursor: pointer; font-size: 14px; }}
  button:hover {{ background: #374151; }}
  .error {{ color: #b91c1c; background: #fef2f2; border-radius: 6px; padding: 12px; }}
  a.download {{ display: inline-block; margin-top: 4px; background: #2563eb; color: #fff; padding: 10px 20px; border-radius: 6px; text-decoration: none; }}
  a.btn-turnover {{ display: inline-block; margin-top: 4px; margin-left: 10px; background: #059669; color: #fff; padding: 10px 20px; border-radius: 6px; text-decoration: none; }}
  label {{ display: block; margin-top: 8px; font-size: 14px; color: #4b5563; }}
  .meta {{ color: #6b7280; font-size: 13px; margin-bottom: 8px; }}
  .stat-grid {{ display: flex; flex-wrap: wrap; gap: 12px; }}
  .stat-box {{ flex: 1; min-width: 140px; border: 1px solid #e5e7eb; border-radius: 8px; padding: 14px 16px; background: #f9fafb; }}
  .stat-box .num {{ font-size: 24px; font-weight: 700; }}
  .stat-box .label {{ font-size: 13px; color: #6b7280; margin-top: 2px; }}
  table {{ border-collapse: collapse; width: 100%; font-size: 13px; margin-top: 8px; }}
  th, td {{ border: 1px solid #e5e7eb; padding: 6px 10px; text-align: left; white-space: nowrap; }}
  th {{ background: #1f2937; color: #fff; }}
  .empty-note {{ color: #9ca3af; font-size: 13px; padding: 8px 0; }}
  .more-note {{ color: #9ca3af; font-size: 12px; margin-top: 6px; }}
</style>
</head>
<body>
  <div class="layout">
    <div class="sidebar">
      <h3>历史整理记录</h3>
      {sidebar_links}
    </div>
    <div class="main">
      <h1>库存整理工具</h1>
      <div class="card">
        <form action="/upload" method="post" enctype="multipart/form-data">
          <label>选这次的库存快照xlsx文件：</label>
          <input type="file" name="file" accept=".xlsx" required>
          <label>如果文件本身不带库存日期（比如260624原始数据.xlsx这种单Sheet1格式），手动填一下日期（可不填，默认用文件修改日期）：</label>
          <input type="date" name="snapshot_date">
          <br><br>
          <button type="submit">上传并整理</button>
        </form>
      </div>
      {result_block}
    </div>
  </div>
</body>
</html>
"""

TURNOVER_PAGE_TEMPLATE = """<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>库存周转率 {date}</title>
<style>
  * {{ box-sizing: border-box; }}
  body {{ font-family: -apple-system, "Microsoft YaHei", sans-serif; margin: 0; color: #1f2937; background: #fafafa; }}
  .header {{ background: #1f2937; color: #fff; padding: 16px 24px; display: flex; align-items: center; gap: 16px; flex-wrap: wrap; }}
  .header h1 {{ margin: 0; font-size: 18px; flex: 1; }}
  .header .meta {{ font-size: 13px; color: #9ca3af; }}
  a.back {{ color: #9ca3af; text-decoration: none; font-size: 13px; }}
  a.back:hover {{ color: #fff; }}
  a.dl {{ background: #2563eb; color: #fff; padding: 8px 16px; border-radius: 6px; text-decoration: none; font-size: 13px; }}
  .toolbar {{ display: flex; gap: 10px; padding: 16px 24px; background: #fff; border-bottom: 1px solid #e5e7eb; flex-wrap: wrap; align-items: center; }}
  .toolbar input {{ border: 1px solid #d1d5db; border-radius: 6px; padding: 7px 12px; font-size: 13px; width: 240px; }}
  .toolbar select {{ border: 1px solid #d1d5db; border-radius: 6px; padding: 7px 12px; font-size: 13px; background: #fff; }}
  .toolbar .count {{ font-size: 13px; color: #6b7280; margin-left: auto; }}
  .table-wrap {{ padding: 0 24px 32px; overflow-x: auto; }}
  table {{ border-collapse: collapse; width: 100%; font-size: 13px; margin-top: 12px; }}
  th {{ background: #1f2937; color: #fff; padding: 8px 12px; text-align: left; white-space: nowrap; cursor: pointer; user-select: none; }}
  th:hover {{ background: #374151; }}
  th .sort-icon {{ margin-left: 4px; opacity: 0.5; }}
  th.sorted .sort-icon {{ opacity: 1; }}
  td {{ border-bottom: 1px solid #e5e7eb; padding: 7px 12px; white-space: nowrap; }}
  tr:hover td {{ background: #f9fafb; }}
  .tag {{ display: inline-block; padding: 2px 8px; border-radius: 3px; font-size: 12px; font-weight: 600; }}
  .no-data {{ text-align: center; color: #9ca3af; padding: 40px; font-size: 14px; }}
</style>
</head>
<body>
<div class="header">
  <a class="back" href="/">← 返回首页</a>
  <h1>库存周转率分析</h1>
  <span class="meta">快照日期：{date} · 周转数据月份：{data_month} · 共 <span id="total-count">{total}</span> 个商品</span>
  <a class="dl" href="/download/turnover/{date}">下载 Excel</a>
</div>
<div class="toolbar">
  <input type="text" id="search" placeholder="搜索商品名称 / 编号…" oninput="applyFilter()">
  <select id="cat-filter" onchange="applyFilter()">
    <option value="">全部分类</option>
    {cat_options}
  </select>
  <select id="status-filter" onchange="applyFilter()">
    <option value="">全部状态</option>
    <option value="快速">快速（≤1月）</option>
    <option value="正常">正常（1-3月）</option>
    <option value="偏慢">偏慢（3-6月）</option>
    <option value="积压">积压（>6月）</option>
    <option value="未知">未知</option>
  </select>
  <span class="count">显示 <span id="shown-count">{total}</span> / {total} 条</span>
</div>
<div class="table-wrap">
  <table id="tv-table">
    <thead>
      <tr>
        <th onclick="sortTable(0)" data-col="0">货品名称 <span class="sort-icon">↕</span></th>
        <th onclick="sortTable(1)" data-col="1">货品编号 <span class="sort-icon">↕</span></th>
        <th onclick="sortTable(2)" data-col="2">页面分类 <span class="sort-icon">↕</span></th>
        <th onclick="sortTable(3)" data-col="3">当前库存 <span class="sort-icon">↕</span></th>
        <th onclick="sortTable(4)" data-col="4">周转月数 <span class="sort-icon">↕</span></th>
        <th onclick="sortTable(5)" data-col="5">库存状态 <span class="sort-icon">↕</span></th>
      </tr>
    </thead>
    <tbody id="tv-body">
      {rows}
    </tbody>
  </table>
  <div id="no-data" class="no-data" style="display:none">没有符合条件的记录</div>
</div>
<script>
const STATUS_COLORS = {status_colors_json};
const ALL_ROWS = Array.from(document.querySelectorAll('#tv-body tr'));
let sortCol = 4, sortAsc = true;

function applyFilter() {{
  const q = document.getElementById('search').value.toLowerCase();
  const catFilter = document.getElementById('cat-filter').value;
  const status = document.getElementById('status-filter').value;
  let shown = 0;
  ALL_ROWS.forEach(tr => {{
    const cells = tr.querySelectorAll('td');
    const name = cells[0].textContent.toLowerCase();
    const sku = cells[1].textContent.toLowerCase();
    const s = cells[5].textContent;
    const cat = cells[2].textContent;
    const match = (!q || name.includes(q) || sku.includes(q))
                && (!catFilter || cat === catFilter)
                && (!status || s === status);
    tr.style.display = match ? '' : 'none';
    if (match) shown++;
  }});
  document.getElementById('shown-count').textContent = shown;
  document.getElementById('no-data').style.display = shown === 0 ? '' : 'none';
}}

function sortTable(col) {{
  if (sortCol === col) {{ sortAsc = !sortAsc; }} else {{ sortCol = col; sortAsc = true; }}
  document.querySelectorAll('th').forEach((th, i) => {{
    th.classList.toggle('sorted', i === col);
    if (i === col) th.querySelector('.sort-icon').textContent = sortAsc ? '↑' : '↓';
    else th.querySelector('.sort-icon').textContent = '↕';
  }});
  const tbody = document.getElementById('tv-body');
  const rows = Array.from(tbody.querySelectorAll('tr'));
  rows.sort((a, b) => {{
    let av = a.querySelectorAll('td')[col].dataset.val || a.querySelectorAll('td')[col].textContent;
    let bv = b.querySelectorAll('td')[col].dataset.val || b.querySelectorAll('td')[col].textContent;
    const an = parseFloat(av), bn = parseFloat(bv);
    if (!isNaN(an) && !isNaN(bn)) return sortAsc ? an - bn : bn - an;
    return sortAsc ? av.localeCompare(bv, 'zh') : bv.localeCompare(av, 'zh');
  }});
  rows.forEach(r => tbody.appendChild(r));
}}
</script>
</body>
</html>
"""


def list_history_dates() -> list[str]:
    if not HISTORY_DIR.exists():
        return []
    dates = [p.stem for p in HISTORY_DIR.glob("*.html")]
    return sorted(dates, reverse=True)


def build_sidebar(active_date: str | None = None) -> str:
    dates = list_history_dates()
    parts = []

    # 月度销量统计——大按钮
    parts.append(
        '<div style="padding:12px 12px 8px">'
        '<a href="/sales" style="display:block;background:#059669;color:#fff;text-decoration:none;'
        'border-radius:8px;padding:14px 16px;font-size:15px;font-weight:700;text-align:center;'
        'letter-spacing:0.5px;box-shadow:0 2px 6px rgba(0,0,0,0.3)">'
        '📊 月度销量统计'
        '</a></div>'
    )

    # 库存整理记录——JS 折叠区
    parts.append(
        '<div style="padding:0 0 4px">'
        '<div onclick="toggleInventory()" id="inv-toggle" style="padding:8px 16px 6px;font-size:12px;'
        'color:#9ca3af;text-transform:uppercase;letter-spacing:0.5px;cursor:pointer;'
        'display:flex;align-items:center;justify-content:space-between">'
        '<span>库存整理记录</span>'
        '<span id="inv-arrow" style="font-size:10px;opacity:0.6">▾</span>'
        '</div>'
        '<div id="inv-list">'
    )
    if not dates:
        parts.append('<div class="empty">还没有整理记录</div>')
    else:
        for d in dates:
            cls = ' class="active"' if d == active_date else ""
            parts.append(f'<a href="/history/{d}"{cls}>{d}</a>')
    parts.append(
        '</div></div>'
        '<script>'
        'function toggleInventory(){'
        'var l=document.getElementById("inv-list");'
        'var a=document.getElementById("inv-arrow");'
        'if(l.style.display==="none"){l.style.display="";a.textContent="▾";}'
        'else{l.style.display="none";a.textContent="▸";}'
        '}'
        '</script>'
    )

    return "".join(parts)


def _render(result_block: str = "", active_date: str | None = None) -> HTMLResponse:
    return HTMLResponse(
        PAGE_TEMPLATE.format(result_block=result_block, sidebar_links=build_sidebar(active_date))
    )


def _esc(v) -> str:
    if pd.isna(v):
        return ""
    return html.escape(str(v))


def _table_html(df: pd.DataFrame, columns: list[str], max_rows: int = 12, status_col: str | None = None) -> str:
    if df.empty:
        return '<div class="empty-note">无数据</div>'
    shown = df.head(max_rows)
    head = "".join(f"<th>{_esc(c)}</th>" for c in columns)
    rows = []
    for _, r in shown.iterrows():
        cells = []
        for c in columns:
            val = r.get(c, "")
            style = ""
            if status_col and c == status_col:
                color = STATUS_COLORS.get(val)
                if color:
                    style = f' style="background:#{color}"'
            cells.append(f"<td{style}>{_esc(val)}</td>")
        rows.append("<tr>" + "".join(cells) + "</tr>")
    more = ""
    if len(df) > max_rows:
        more = f'<div class="more-note">共 {len(df)} 条，仅展示前 {max_rows} 条，完整结果请下载Excel</div>'
    return f"<table><thead><tr>{head}</tr></thead><tbody>{''.join(rows)}</tbody></table>{more}"


QTY_BRANDS = [("STTOKE", "STTOKE"), ("慕咖", "慕咖"), ("食品", "巴恩天然")]


def _build_turnover_html(goods: pd.DataFrame) -> str:
    """结果摘要页里的周转率小块（品牌均值 + 积压列表），详细页跳转到 /turnover/{date}"""
    if "最新月周转月数" not in goods.columns or goods["最新月周转月数"].isna().all():
        return ""
    has_data = goods[goods["最新月周转月数"].notna()]
    if has_data.empty:
        return ""

    data_month = has_data["周转数据月份"].iloc[0] if "周转数据月份" in has_data.columns else ""
    brand_rows = []
    for brand, g in has_data.groupby("品牌"):
        avg_t = round(g["最新月周转月数"].mean())
        median_t = round(g["最新月周转月数"].median())
        counts = g["库存状态"].value_counts()
        status_str = "  ".join(
            f'<span style="background:#{STATUS_COLORS.get(s,"")}; padding:2px 6px; border-radius:3px">'
            f'{_esc(s)} {int(counts.get(s, 0))}</span>'
            for s in ["快速", "正常", "偏慢", "积压"] if counts.get(s, 0) > 0
        )
        brand_rows.append(
            f"<tr><td>{_esc(brand)}</td><td>{avg_t}</td><td>{median_t}</td><td>{status_str}</td></tr>"
        )
    brand_table = (
        "<table><thead><tr><th>品牌</th><th>平均周转月数</th><th>中位周转月数</th><th>状态分布</th></tr></thead>"
        f"<tbody>{''.join(brand_rows)}</tbody></table>"
    )
    month_note = f"（数据月份：{_esc(data_month)}）" if data_month else ""
    return f"<h2>库存周转率概览 {month_note}</h2>{brand_table}"


def build_result_html(tables: dict, entry: tuple | None) -> str:
    goods = tables["goods"]
    packaging = tables["packaging"]
    snapshot_date = tables["snapshot_date"]

    qty_boxes = ""
    if "汇总" in goods.columns:
        for brand_key, label in QTY_BRANDS:
            total_qty = int(goods.loc[goods["品牌"] == brand_key, "汇总"].sum())
            qty_boxes += (
                f'<div class="stat-box"><div class="num">{total_qty:,}</div>'
                f'<div class="label">{_esc(label)} 总库存数量</div></div>'
            )

    goods_counts = goods["品牌"].value_counts()
    pkg_counts = packaging["品牌"].value_counts()

    stat_boxes = "".join(
        f'<div class="stat-box"><div class="num">{v}</div><div class="label">{_esc(k)} 商品</div></div>'
        for k, v in goods_counts.items()
    )
    stat_boxes += "".join(
        f'<div class="stat-box"><div class="num">{v}</div><div class="label">{_esc(k)} 包装物料</div></div>'
        for k, v in pkg_counts.items()
    )

    status_pivot = (
        goods[goods["库存状态"].notna()]
        .groupby(["品牌", "库存状态"])
        .size()
        .unstack(fill_value=0)
        if "库存状态" in goods.columns and goods["库存状态"].notna().any()
        else pd.DataFrame()
    )
    status_cols = [c for c in ["快速", "正常", "偏慢", "积压", "未知"] if c in status_pivot.columns]
    status_table = ""
    if not status_pivot.empty:
        rows = []
        head = "<th>品牌</th>" + "".join(f"<th>{c}</th>" for c in status_cols)
        for brand, row in status_pivot.iterrows():
            cells = f"<td>{_esc(brand)}</td>"
            for c in status_cols:
                bg = STATUS_COLORS.get(c, "")
                cells += f'<td style="background:#{bg}">{int(row[c])}</td>'
            rows.append(f"<tr>{cells}</tr>")
        status_table = f"<table><thead><tr>{head}</tr></thead><tbody>{''.join(rows)}</tbody></table>"
    else:
        status_table = '<div class="empty-note">无周转状态数据</div>'

    expiring = tables.get("expiring_soon", pd.DataFrame())
    expiring_cols = [c for c in ["货品名称", "货品编号", "汇总", "批次号", "过期日期", "距离到期天数_最近批次"] if c in expiring.columns]
    expiring_html = _table_html(expiring, expiring_cols, max_rows=12)

    low_stock = tables.get("low_stock", pd.DataFrame())
    low_stock_html = ""
    if low_stock is not None and len(low_stock):
        low_cols = [c for c in low_stock.columns if c in ("货品名称", "货品编号", "品牌", "汇总")]
        low_stock_html = f"""
        <h2>低库存预警（汇总≤10）</h2>
        {_table_html(low_stock, low_cols, max_rows=12)}
        """

    turnover_summary = _build_turnover_html(goods)

    analyzed_at = datetime.now().strftime("%Y-%m-%d %H:%M")
    meta_line = f"快照日期：{snapshot_date}（格式：{tables['format']}）· 分析时间：{analyzed_at}"
    if entry:
        meta_line = f"刚登记的快照：{_esc(entry[2])}（日期 {entry[0]}） · " + meta_line

    download_name = f"库存整理_{snapshot_date}.xlsx"

    return f"""
    <div class="card">
      <div class="meta">{meta_line}</div>
      <a class="download" href="/download/{snapshot_date}?name={download_name}">下载完整整理结果（含所有品牌sheet）</a>
      <a class="btn-turnover" href="/turnover/{snapshot_date}">查看库存周转率详细页 →</a>

      <h2>总库存数量统计</h2>
      <div class="stat-grid">{qty_boxes}</div>

      <h2>品牌数量概览</h2>
      <div class="stat-grid">{stat_boxes}</div>

      <h2>库存状态分布（按品牌）</h2>
      {status_table}

      {turnover_summary}

      <h2>临期预警（1年内到期，按紧迫度排序）</h2>
      {expiring_html}

      {low_stock_html}
    </div>
    """


def _save_turnover_json(snapshot_date: str, goods: pd.DataFrame) -> None:
    """把全部商品存成 JSON，供周转率详情页使用（含无周转数据的商品，显示—）"""
    cols = [c for c in ["货品名称", "货品编号", "类型", "品牌", "汇总", "最新月周转月数", "库存状态", "周转数据月份"] if c in goods.columns]
    tv = goods[cols].copy()
    tv["最新月周转月数"] = tv["最新月周转月数"].apply(lambda x: int(round(x)) if pd.notna(x) else None)
    tv["页面分类"] = tv["类型"] + "-" + tv["品牌"]
    HISTORY_DIR.mkdir(parents=True, exist_ok=True)
    (HISTORY_DIR / f"{snapshot_date}_turnover.json").write_text(
        tv.to_json(orient="records", force_ascii=False), encoding="utf-8"
    )


def save_history(snapshot_date: str, result_html: str, goods: pd.DataFrame) -> None:
    HISTORY_DIR.mkdir(parents=True, exist_ok=True)
    (HISTORY_DIR / f"{snapshot_date}.html").write_text(result_html, encoding="utf-8")
    shutil.copy(clean_latest.OUT_PATH, HISTORY_DIR / f"{snapshot_date}.xlsx")
    _save_turnover_json(snapshot_date, goods)


def _build_turnover_page(snapshot_date: str) -> HTMLResponse:
    path = HISTORY_DIR / f"{snapshot_date}_turnover.json"
    if not path.exists():
        raise HTTPException(404, "该期没有周转率数据，请重新上传快照文件")

    records = json.loads(path.read_text(encoding="utf-8"))
    if not records:
        raise HTTPException(404, "周转率数据为空")

    data_month = records[0].get("周转数据月份", "") if records else ""
    # 按页面分类排序：商品类在前，包装物料在后；同类内按品牌顺序
    CAT_ORDER = ["商品-STTOKE", "商品-慕咖", "商品-食品", "商品-周边",
                 "包装物料-STTOKE", "包装物料-慕咖", "包装物料-食品", "包装物料-未分类"]
    cats_in_data = {r.get("页面分类", "") for r in records if r.get("页面分类")}
    ordered_cats = [c for c in CAT_ORDER if c in cats_in_data] + sorted(cats_in_data - set(CAT_ORDER))
    cat_options = "".join(f'<option value="{html.escape(c)}">{html.escape(c)}</option>' for c in ordered_cats)

    status_colors_json = json.dumps(STATUS_COLORS, ensure_ascii=False)

    def _status_tag(s):
        color = STATUS_COLORS.get(s, "e5e7eb")
        return f'<span class="tag" style="background:#{color}">{html.escape(str(s))}</span>'

    rows_html = []
    for r in records:
        name = html.escape(str(r.get("货品名称", "")))
        sku = html.escape(str(r.get("货品编号", "")))
        cat = html.escape(str(r.get("页面分类", "")))
        qty = r.get("汇总", 0) or 0
        tm = r.get("最新月周转月数")
        tm_display = str(tm) if tm is not None else "—"
        tm_val = str(tm) if tm is not None else "9999"
        status = str(r.get("库存状态", ""))
        rows_html.append(
            f'<tr>'
            f'<td>{name}</td>'
            f'<td>{sku}</td>'
            f'<td>{cat}</td>'
            f'<td data-val="{qty}">{qty:,}</td>'
            f'<td data-val="{tm_val}">{tm_display}</td>'
            f'<td>{_status_tag(status)}</td>'
            f'</tr>'
        )

    page = TURNOVER_PAGE_TEMPLATE.format(
        date=snapshot_date,
        data_month=html.escape(str(data_month)),
        total=len(records),
        cat_options=cat_options,
        rows="".join(rows_html),
        status_colors_json=status_colors_json,
    )
    return HTMLResponse(page)


def _build_turnover_excel(snapshot_date: str) -> StreamingResponse:
    path = HISTORY_DIR / f"{snapshot_date}_turnover.json"
    if not path.exists():
        raise HTTPException(404, "该期没有周转率数据")

    records = json.loads(path.read_text(encoding="utf-8"))
    HEADER_FILL = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    HEADER_FONT = Font(color="FFFFFF", bold=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "库存周转率"

    headers = ["货品名称", "货品编号", "页面分类", "品牌", "当前库存", "周转月数", "库存状态", "周转数据月份"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center")

    status_fills = {s: PatternFill(start_color=color, end_color=color, fill_type="solid")
                    for s, color in STATUS_COLORS.items()}

    for row_idx, r in enumerate(records, 2):
        status = str(r.get("库存状态", ""))
        tm = r.get("最新月周转月数")
        vals = [
            r.get("货品名称", ""),
            r.get("货品编号", ""),
            r.get("页面分类", ""),
            r.get("品牌", ""),
            r.get("汇总", 0),
            tm,
            status,
            r.get("周转数据月份", ""),
        ]
        for col, v in enumerate(vals, 1):
            cell = ws.cell(row=row_idx, column=col, value=v)
            if col == 6 and status in status_fills:
                cell.fill = status_fills[status]

    # 自动列宽
    for col in range(1, len(headers) + 1):
        from openpyxl.utils import get_column_letter
        letter = get_column_letter(col)
        max_len = max(
            (len(str(ws.cell(row=r, column=col).value or "")) for r in range(1, ws.max_row + 1)),
            default=10,
        )
        ws.column_dimensions[letter].width = min(max(max_len + 2, 10), 50)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"库存周转率_{snapshot_date}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.get("/", response_class=HTMLResponse)
def index():
    dates = list_history_dates()
    if not dates:
        return _render()
    latest = dates[0]
    # 实时渲染最新期（只跑 clean，不重算周转率/不重写Excel，速度快）
    # 保证代码更新后新功能立刻出现，不依赖缓存 HTML
    try:
        tables = clean_latest.clean()
        result_html = build_result_html(tables, None)
        save_history(tables["snapshot_date"], result_html, tables["goods"])
        return _render(result_html, active_date=tables["snapshot_date"])
    except Exception:
        result_html = (HISTORY_DIR / f"{latest}.html").read_text(encoding="utf-8")
        return _render(result_html, active_date=latest)


@app.get("/history/{date}", response_class=HTMLResponse)
def history(date: str):
    path = HISTORY_DIR / f"{date}.html"
    if not path.exists():
        raise HTTPException(404, "没有这一期的记录")
    return _render(path.read_text(encoding="utf-8"), active_date=date)


@app.get("/turnover/{date}")
def turnover_page(date: str):
    return _build_turnover_page(date)


@app.get("/download/turnover/{date}")
def download_turnover(date: str):
    return _build_turnover_excel(date)


@app.post("/upload", response_class=HTMLResponse)
async def upload(file: UploadFile = File(...), snapshot_date: str = Form(default="")):
    if not file.filename.lower().endswith(".xlsx"):
        return _render('<div class="error">只支持 .xlsx 文件</div>')

    RAW_DIR.mkdir(parents=True, exist_ok=True)
    dest_path = RAW_DIR / file.filename
    with open(dest_path, "wb") as f:
        shutil.copyfileobj(file.file, f)

    try:
        entry = clean_latest.register_snapshot(dest_path, date_override=snapshot_date or None)
        tables = clean_latest.run_pipeline()
    except Exception as e:
        dest_path.unlink(missing_ok=True)
        return _render(f'<div class="error">处理失败：{html.escape(str(e))}</div>')

    result_html = build_result_html(tables, entry)
    save_history(tables["snapshot_date"], result_html, tables["goods"])

    return _render(result_html, active_date=tables["snapshot_date"])


@app.get("/download/{date}")
def download(date: str, name: str = "库存整理.xlsx"):
    path = HISTORY_DIR / f"{date}.xlsx"
    if not path.exists():
        raise HTTPException(404, "文件不存在")
    return FileResponse(path, filename=name, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ── 月度销量页面 ────────────────────────────────────────────

@app.post("/upload/sales", response_class=HTMLResponse)
async def upload_sales(files: list[UploadFile] = File(...)):
    SALES_DIR.mkdir(parents=True, exist_ok=True)
    saved = []
    errors = []
    for file in files:
        name = file.filename
        if not name.endswith(".json"):
            errors.append(f"{name}：只支持 .json 文件")
            continue
        if not name.startswith("results_") or len(name) != len("results_202606.json"):
            errors.append(f"{name}：文件名格式应为 results_YYYYMM.json")
            continue
        dest = SALES_DIR / name
        with open(dest, "wb") as f:
            shutil.copyfileobj(file.file, f)
        saved.append(name)
    from fastapi.responses import RedirectResponse
    return RedirectResponse(url="/sales", status_code=303)


@app.get("/sales", response_class=HTMLResponse)
def sales_page():
    try:
        sales, names = load_monthly_sales()
        months = get_sales_months()
    except Exception as e:
        return HTMLResponse(f"<p>读取销量数据失败：{html.escape(str(e))}</p>")

    # 已上传的文件列表
    uploaded_files = sorted(SALES_DIR.glob("results_??????.json")) if SALES_DIR.exists() else []
    upload_list = "".join(f"<li>{f.name}</li>" for f in uploaded_files) or "<li style='color:#9ca3af'>暂无文件</li>"

    if not sales:
        return HTMLResponse(f"""<!DOCTYPE html><html lang="zh-CN"><head><meta charset="utf-8"><title>月度销量统计</title>
        <style>body{{font-family:-apple-system,"Microsoft YaHei",sans-serif;margin:0;background:#fafafa}}
        .header{{background:#1f2937;color:#fff;padding:16px 24px;display:flex;align-items:center;gap:16px}}
        .header h1{{margin:0;font-size:18px}} a.back{{color:#9ca3af;text-decoration:none;font-size:13px}}
        .tip{{padding:32px 24px;color:#6b7280}}</style></head><body>
        <div class="header"><a class="back" href="/">← 返回首页</a><h1>月度销量统计</h1></div>
        <div style="background:#f0fdf4;border-bottom:1px solid #bbf7d0;padding:12px 24px;display:flex;align-items:center;gap:16px;flex-wrap:wrap">
          <form action="/upload/sales" method="post" enctype="multipart/form-data" style="display:flex;align-items:center;gap:8px;flex-wrap:wrap">
            <span style="font-size:13px;color:#166534;font-weight:600">上传销量数据：</span>
            <input type="file" name="files" accept=".json" multiple style="font-size:13px">
            <button type="submit" style="background:#16a34a;color:#fff;border:none;padding:6px 14px;border-radius:6px;cursor:pointer;font-size:13px">上传</button>
          </form>
        </div>
        <div class="tip">还没有销量数据，请上传 results_YYYYMM.json 文件（来自 sales_agent 项目的 data 目录）。</div>
        </body></html>""")

    # 从库存注册表补充商品名（优先用库存系统里的名字）
    try:
        name_reg = clean_latest.load_name_registry()
    except Exception:
        name_reg = {}

    # 构建行数据：sku, name, {month: qty}
    rows_data = []
    for sku, mdata in sorted(sales.items()):
        name = name_reg.get(sku) or names.get(sku, sku)
        total = sum(mdata.values())
        rows_data.append({"sku": sku, "name": name, "months": mdata, "total": total})

    # 按总销量降序
    rows_data.sort(key=lambda r: r["total"], reverse=True)

    # 表格 HTML
    month_headers = "".join(f"<th onclick=\"sortTable({i+2})\" data-col=\"{i+2}\">{m} <span class=\"sort-icon\">↕</span></th>" for i, m in enumerate(months))
    month_headers += f"<th onclick=\"sortTable({len(months)+2})\" data-col=\"{len(months)+2}\">合计 <span class=\"sort-icon\">↓</span></th>"

    tbody_rows = []
    for r in rows_data:
        cells = f'<td>{html.escape(r["name"])}</td><td>{html.escape(r["sku"])}</td>'
        row_total = 0
        for m in months:
            q = r["months"].get(m, 0)
            row_total += q
            cells += f'<td data-val="{q}">{q if q else ""}</td>'
        cells += f'<td data-val="{row_total}"><strong>{row_total}</strong></td>'
        tbody_rows.append(f"<tr>{cells}</tr>")

    month_col_count = len(months) + 3  # 名称+编号+各月+合计

    page = f"""<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>月度销量统计</title>
<style>
  * {{ box-sizing: border-box; }}
  body {{ font-family: -apple-system, "Microsoft YaHei", sans-serif; margin: 0; color: #1f2937; background: #fafafa; }}
  .header {{ background: #1f2937; color: #fff; padding: 16px 24px; display: flex; align-items: center; gap: 16px; flex-wrap: wrap; }}
  .header h1 {{ margin: 0; font-size: 18px; flex: 1; }}
  .header .meta {{ font-size: 13px; color: #9ca3af; }}
  a.back {{ color: #9ca3af; text-decoration: none; font-size: 13px; }}
  a.back:hover {{ color: #fff; }}
  a.dl {{ background: #2563eb; color: #fff; padding: 8px 16px; border-radius: 6px; text-decoration: none; font-size: 13px; }}
  .toolbar {{ display: flex; gap: 10px; padding: 16px 24px; background: #fff; border-bottom: 1px solid #e5e7eb; flex-wrap: wrap; align-items: center; }}
  .toolbar input {{ border: 1px solid #d1d5db; border-radius: 6px; padding: 7px 12px; font-size: 13px; width: 240px; }}
  .toolbar .count {{ font-size: 13px; color: #6b7280; margin-left: auto; }}
  .table-wrap {{ padding: 0 24px 32px; overflow-x: auto; }}
  table {{ border-collapse: collapse; width: 100%; font-size: 13px; margin-top: 12px; }}
  th {{ background: #1f2937; color: #fff; padding: 8px 12px; text-align: right; white-space: nowrap; cursor: pointer; user-select: none; }}
  th:first-child, th:nth-child(2) {{ text-align: left; }}
  th:hover {{ background: #374151; }}
  th .sort-icon {{ margin-left: 4px; opacity: 0.5; }}
  th.sorted .sort-icon {{ opacity: 1; }}
  td {{ border-bottom: 1px solid #e5e7eb; padding: 7px 12px; white-space: nowrap; text-align: right; }}
  td:first-child, td:nth-child(2) {{ text-align: left; }}
  tr:hover td {{ background: #f9fafb; }}
  .no-data {{ text-align: center; color: #9ca3af; padding: 40px; font-size: 14px; }}
</style>
</head>
<body>
<div class="header">
  <a class="back" href="/">← 返回首页</a>
  <h1>月度销量统计（各平台汇总）</h1>
  <span class="meta">数据月份：{months[0] if months else ""} ~ {months[-1] if months else ""}  · 共 <span id="total-count">{len(rows_data)}</span> 个 SKU</span>
  <a class="dl" href="/download/sales">下载 Excel</a>
</div>
<div style="background:#f0fdf4;border-bottom:1px solid #bbf7d0;padding:12px 24px;display:flex;align-items:center;gap:16px;flex-wrap:wrap">
  <form action="/upload/sales" method="post" enctype="multipart/form-data" style="display:flex;align-items:center;gap:8px;flex-wrap:wrap">
    <span style="font-size:13px;color:#166534;font-weight:600">上传销量数据：</span>
    <input type="file" name="files" accept=".json" multiple style="font-size:13px">
    <button type="submit" style="background:#16a34a;color:#fff;border:none;padding:6px 14px;border-radius:6px;cursor:pointer;font-size:13px">上传</button>
  </form>
  <details style="font-size:12px;color:#166534;cursor:pointer">
    <summary>已上传文件（{len(uploaded_files)} 个）</summary>
    <ul style="margin:4px 0 0 16px;padding:0">{upload_list}</ul>
  </details>
</div>
<div class="toolbar">
  <input type="text" id="search" placeholder="搜索商品名称 / SKU编码…" oninput="applyFilter()">
  <span class="count">显示 <span id="shown-count">{len(rows_data)}</span> / {len(rows_data)} 条</span>
</div>
<div class="table-wrap">
  <table id="sales-table">
    <thead>
      <tr>
        <th onclick="sortTable(0)" data-col="0">商品名称 <span class="sort-icon">↕</span></th>
        <th onclick="sortTable(1)" data-col="1">SKU编码 <span class="sort-icon">↕</span></th>
        {month_headers}
      </tr>
    </thead>
    <tbody id="sales-body">
      {"".join(tbody_rows)}
    </tbody>
  </table>
  <div id="no-data" class="no-data" style="display:none">没有符合条件的记录</div>
</div>
<script>
const ALL_ROWS = Array.from(document.querySelectorAll('#sales-body tr'));
let sortCol = {len(months) + 2}, sortAsc = false;

function applyFilter() {{
  const q = document.getElementById('search').value.toLowerCase();
  let shown = 0;
  ALL_ROWS.forEach(tr => {{
    const cells = tr.querySelectorAll('td');
    const name = cells[0].textContent.toLowerCase();
    const sku = cells[1].textContent.toLowerCase();
    const match = !q || name.includes(q) || sku.includes(q);
    tr.style.display = match ? '' : 'none';
    if (match) shown++;
  }});
  document.getElementById('shown-count').textContent = shown;
  document.getElementById('no-data').style.display = shown === 0 ? '' : 'none';
}}

function sortTable(col) {{
  if (sortCol === col) {{ sortAsc = !sortAsc; }} else {{ sortCol = col; sortAsc = true; }}
  document.querySelectorAll('th').forEach((th, i) => {{
    th.classList.toggle('sorted', i === col);
    if (i === col) th.querySelector('.sort-icon').textContent = sortAsc ? '↑' : '↓';
    else th.querySelector('.sort-icon').textContent = '↕';
  }});
  const tbody = document.getElementById('sales-body');
  const rows = Array.from(tbody.querySelectorAll('tr'));
  rows.sort((a, b) => {{
    let av = a.querySelectorAll('td')[col].dataset.val || a.querySelectorAll('td')[col].textContent;
    let bv = b.querySelectorAll('td')[col].dataset.val || b.querySelectorAll('td')[col].textContent;
    const an = parseFloat(av), bn = parseFloat(bv);
    if (!isNaN(an) && !isNaN(bn)) return sortAsc ? an - bn : bn - an;
    return sortAsc ? av.localeCompare(bv, 'zh') : bv.localeCompare(av, 'zh');
  }});
  rows.forEach(r => tbody.appendChild(r));
}}
</script>
</body>
</html>"""
    return HTMLResponse(page)


@app.get("/download/sales")
def download_sales():
    try:
        sales, names = load_monthly_sales()
        months = get_sales_months()
    except Exception as e:
        raise HTTPException(500, str(e))

    try:
        name_reg = clean_latest.load_name_registry()
    except Exception:
        name_reg = {}

    HEADER_FILL = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    HEADER_FONT = Font(color="FFFFFF", bold=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "月度销量"

    headers = ["商品名称", "SKU编码"] + months + ["合计"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center")

    rows_data = []
    for sku, mdata in sales.items():
        name = name_reg.get(sku) or names.get(sku, sku)
        rows_data.append((name, sku, mdata, sum(mdata.values())))
    rows_data.sort(key=lambda r: r[3], reverse=True)

    for row_idx, (name, sku, mdata, total) in enumerate(rows_data, 2):
        ws.cell(row=row_idx, column=1, value=name)
        ws.cell(row=row_idx, column=2, value=sku)
        for col_idx, m in enumerate(months, 3):
            ws.cell(row=row_idx, column=col_idx, value=mdata.get(m, 0) or None)
        ws.cell(row=row_idx, column=len(months) + 3, value=total).font = Font(bold=True)

    from openpyxl.utils import get_column_letter
    for col in range(1, len(headers) + 1):
        letter = get_column_letter(col)
        max_len = max(
            (len(str(ws.cell(row=r, column=col).value or "")) for r in range(1, ws.max_row + 1)),
            default=8,
        )
        ws.column_dimensions[letter].width = min(max(max_len + 2, 8), 50)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=\"月度销量统计.xlsx\""},
    )
