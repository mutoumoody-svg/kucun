"""
整理 260624原始数据.xlsx（当前库存盘点导出，6/24，单Sheet1，按仓库逐行展开）。

跟标准周报/ERP格式都不同，此文件里410行混杂着：
- 真实可销售商品（STTOKE/慕咖随行杯、巴恩天然蜂蜜苹果醋、Alpaka/Hario/Delter等周边）
- 包装物料/辅料（标签、贴纸、外盒、纸箱、腰封、胶条、磨砂袋、展示架、合格证、宣传册等）

保质期信息来源：260624原始数据.xlsx本身不含保质期字段，只有ERP批次导出
（如 副本库存6.1(1).xlsx）才记录批次号/生产日期/过期日期/距离到期天数，且只有
食品品牌（巴恩天然蜂蜜/苹果醋/康蜜乐等）才有真实批次记录，其余行的日期都是
"0000-00-00"占位符，代表未做保质期管理。所以保质期数据按SKU(商家编码=货品编号)
从最近一次ERP批次导出里join过来，并标注数据来源日期，避免和6/24当前库存日期混淆。

整理逻辑：
1. 同一SKU不同仓库的库存量横向展开成并列的列（一行一个产品），后面加一列"汇总"，
   不再按仓库拆成多行——这样一眼就能看出总库存是多少
2. 按品牌分类（复用 categorize.py）：STTOKE/慕咖/食品/周边商品，不区分商品种类
   —— 杯子、蜂蜜、苹果醋、咖啡只要是同一品牌就放在同一张表里，不按食品/非食品拆开
3. 包装物料同样按品牌单独分表，不与对应品牌的商品混在一张表里
4. 食品品牌的商品附保质期信息（如有批次记录）
5. 出库数量是"当月累计出库"（月初到本期快照的累计下降量，中途补货不会冲淡这个数），
   不是单纯跟上一期快照的差值
6. 货品名称放在第一列，同规格关键词（如12OZ/16OZ）的产品排在一起
"""
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from categorize import (
    BRAND_ORDER,
    MOODY_TOP_ITEMS,
    SKU_ALIAS,
    STATUS_COLORS,
    classify_brand,
    classify_material,
    extract_group_keyword,
    food_sort_key,
    group_sort_key,
    moody_sort_key,
)
from monthly_outbound import cumulative_month_outbound
from turnover_lookup import load_latest_turnover

RAW_DIR = Path(__file__).resolve().parent.parent / "raw_data"
SRC_PATH = RAW_DIR / "260624原始数据.xlsx"
EXPIRY_SRC_PATH = RAW_DIR / "副本库存6.1(1).xlsx"
OUT_PATH = Path(__file__).resolve().parent.parent / "output" / "260624库存整理.xlsx"
SNAPSHOT_DATE = "2026-06-24"

HEADER_FILL = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def _write_header(ws, headers, row=1):
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=col, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center")
    ws.freeze_panes = ws.cell(row=row + 1, column=1)


def _autosize(ws, n_cols):
    for col in range(1, n_cols + 1):
        letter = get_column_letter(col)
        max_len = max(
            (len(str(ws.cell(row=r, column=col).value or "")) for r in range(1, ws.max_row + 1)),
            default=10,
        )
        ws.column_dimensions[letter].width = min(max(max_len + 2, 10), 50)


TOTAL_FONT = Font(bold=True, size=12, color="000000")


def _write_rows(ws, df: pd.DataFrame, columns: list[str], start_row: int, number_format: dict | None = None) -> int:
    """从start_row开始写数据行（不写表头），返回写完之后的下一行行号"""
    status_col_idx = columns.index("库存状态") + 1 if "库存状态" in columns else None
    total_col_idx = columns.index("汇总") + 1 if "汇总" in columns else None
    row = start_row
    for _, r in df.iterrows():
        for j, col in enumerate(columns, start=1):
            cell = ws.cell(row=row, column=j, value=r[col])
            if number_format and col in number_format:
                cell.number_format = number_format[col]
            if status_col_idx and j == status_col_idx:
                color = STATUS_COLORS.get(r["库存状态"])
                if color:
                    cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            if total_col_idx and j == total_col_idx:
                cell.font = TOTAL_FONT
        row += 1
    return row


def _write_df(ws, df: pd.DataFrame, number_format: dict | None = None):
    columns = list(df.columns)
    _write_header(ws, columns)
    _write_rows(ws, df, columns, start_row=2, number_format=number_format)
    _autosize(ws, len(columns))


def _write_df_with_gap(ws, df_top: pd.DataFrame, df_rest: pd.DataFrame, gap_rows: int, number_format: dict | None = None):
    """先写df_top，空gap_rows行，再接着写df_rest，表头只写一次"""
    columns = list(df_top.columns)
    _write_header(ws, columns)
    next_row = _write_rows(ws, df_top, columns, start_row=2, number_format=number_format)
    _write_rows(ws, df_rest, columns, start_row=next_row + gap_rows, number_format=number_format)
    _autosize(ws, len(columns))


# 保质期来源文件按时间顺序列出。同一SKU在多个来源里都出现时，靠后的（更新的）
# 来源优先——这样像"NEWBNMGO100"这种6/1批次快照里还没建档、6/29才有的新货号
# 也能补到保质期，不会因为join到旧文件就拿不到数据。
EXPIRY_SOURCES = [
    (EXPIRY_SRC_PATH, "库存6.1"),
    (RAW_DIR / "副本库存6.29.xlsx", "蜂蜜"),
]


def load_expiry_table() -> pd.DataFrame:
    """汇总所有ERP批次导出里的真实保质期记录（过滤掉0000-00-00占位符），
    同一SKU在多个来源里出现时优先用较新来源的批次。"""
    frames = []
    for path, sheet in EXPIRY_SOURCES:
        if not path.exists():
            continue
        erp = pd.read_excel(path, sheet_name=sheet)
        snapshot_date = str(erp["库存日期"].iloc[0])
        real = erp[erp["生产日期"].astype(str) != "0000-00-00 00:00:00"].copy()
        real["sku"] = real["商家编码"].replace(SKU_ALIAS)
        real["保质期数据日期"] = snapshot_date
        real["来源顺序"] = len(frames)
        frames.append(real)

    if not frames:
        return pd.DataFrame(columns=["sku", "批次号", "生产日期", "过期日期", "距离到期天数_最近批次", "保质期数据日期"])

    combined = pd.concat(frames, ignore_index=True)
    # 同一SKU+同一批次号在多个来源里重复出现时，只保留较新来源的那条
    combined = combined.sort_values("来源顺序").drop_duplicates(subset=["sku", "批次号"], keep="last")

    combined = combined.sort_values(["sku", "距离到期天数"])
    agg = combined.groupby("sku").agg(
        批次号=("批次号", lambda s: " / ".join(str(x) for x in s)),
        生产日期=("生产日期", "first"),
        过期日期=("过期日期", "first"),
        距离到期天数_最近批次=("距离到期天数", "min"),
        保质期数据日期=("保质期数据日期", "last"),
    ).reset_index()
    return agg


def clean() -> dict:
    df = pd.read_excel(SRC_PATH, sheet_name="Sheet1")
    df["货品编号"] = df["货品编号"].replace(SKU_ALIAS)
    warehouses = sorted(df["仓库"].unique())

    # 同一SKU不同仓库横向展开成并列的列，再加一列汇总
    wide = df.pivot_table(
        index=["货品编号", "货品名称"], columns="仓库", values="库存量", aggfunc="sum", fill_value=0
    ).reset_index()
    wide["汇总"] = wide[warehouses].sum(axis=1)

    wide["类型"] = wide["货品名称"].apply(classify_material)
    wide["品牌"] = wide["货品名称"].apply(classify_brand)
    wide["规格分组"] = wide["货品名称"].apply(extract_group_keyword)

    expiry = load_expiry_table()
    wide = wide.merge(expiry, left_on="货品编号", right_on="sku", how="left").drop(columns=["sku"])

    turnover = load_latest_turnover()
    wide = wide.merge(turnover, left_on="货品编号", right_on="sku", how="left").drop(columns=["sku"])

    outbound = cumulative_month_outbound(SNAPSHOT_DATE)
    wide = wide.merge(outbound, left_on="货品编号", right_on="sku", how="left").drop(columns=["sku"])

    def _sort_weight(r) -> int:
        if r["品牌"] == "食品":
            return food_sort_key(r["货品名称"])
        if r["品牌"] == "慕咖":
            if r["货品名称"] in MOODY_TOP_ITEMS:
                return MOODY_TOP_ITEMS.index(r["货品名称"])
            return moody_sort_key(r["货品名称"])
        return group_sort_key(r["规格分组"])

    def reorder(d: pd.DataFrame) -> pd.DataFrame:
        d = d.copy()
        # 食品/慕咖品牌按人工指定的固定顺序排，其余品牌按规格分组排
        d["_排序权重"] = d.apply(_sort_weight, axis=1)
        d = d.sort_values(
            ["_排序权重", "规格分组", "距离到期天数_最近批次", "汇总"],
            ascending=[True, True, True, False],
            na_position="last",
        ).drop(columns=["_排序权重"])
        cols = ["货品名称", "规格分组"] + [c for c in d.columns if c not in ("货品名称", "规格分组")]
        return d[cols]

    goods = reorder(wide[wide["类型"] == "商品"])
    packaging = reorder(wide[wide["类型"] == "包装物料"])

    low_stock = goods[(goods["汇总"] > 0) & (goods["汇总"] <= 10)]
    expiring_soon = goods[
        (goods["品牌"] == "食品") & goods["距离到期天数_最近批次"].notna() & (goods["距离到期天数_最近批次"] <= 365)
    ]

    return {
        "goods": goods,
        "packaging": packaging,
        "low_stock": low_stock,
        "expiring_soon": expiring_soon,
        "warehouses": warehouses,
    }


def _brand_cols(brand: str, warehouses: list[str]) -> list[str]:
    base = ["货品名称", "规格分组", "货品编号"] + warehouses + ["汇总"]
    if brand == "食品":
        base += ["批次号", "生产日期", "过期日期", "距离到期天数_最近批次", "保质期数据日期"]
    base += ["累计出库数量", "出库数据周期", "最新月周转月数", "库存状态", "周转数据月份"]
    return base


def build_workbook(tables: dict):
    wb = Workbook()
    wb.remove(wb.active)
    warehouses = tables["warehouses"]
    qty_fmt = {w: "#,##0" for w in warehouses}
    qty_fmt.update({"汇总": "#,##0", "累计出库数量": "#,##0", "最新月周转月数": "0.0"})
    idx = 1

    for brand in BRAND_ORDER:
        sub = tables["goods"][tables["goods"]["品牌"] == brand]
        if sub.empty:
            continue
        ws = wb.create_sheet(f"{idx:02d}_商品-{brand}")
        cols = _brand_cols(brand, warehouses)
        if brand == "慕咖":
            top = sub[sub["货品名称"].isin(MOODY_TOP_ITEMS)]
            rest = sub[~sub["货品名称"].isin(MOODY_TOP_ITEMS)]
            _write_df_with_gap(ws, top[cols], rest[cols], gap_rows=2, number_format=qty_fmt)
        else:
            _write_df(ws, sub[cols], number_format=qty_fmt)
        idx += 1

    for brand in BRAND_ORDER:
        sub = tables["packaging"][tables["packaging"]["品牌"] == brand]
        if sub.empty:
            continue
        ws = wb.create_sheet(f"{idx:02d}_包装物料-{brand}")
        cols = ["货品名称", "规格分组", "货品编号"] + warehouses + ["汇总"]
        _write_df(ws, sub[cols], number_format=qty_fmt)
        idx += 1

    ws_low = wb.create_sheet(f"{idx:02d}_低库存预警")
    idx += 1
    low = tables["low_stock"].copy()
    low["提醒"] = "总库存≤10，建议核实是否需要补货/下架"
    _write_df(
        ws_low,
        low[["货品名称", "货品编号", "品牌"] + warehouses + ["汇总", "提醒"]],
        number_format=qty_fmt,
    )

    ws_exp = wb.create_sheet(f"{idx:02d}_临期预警")
    expiring = tables["expiring_soon"].copy()
    if len(expiring):
        expiring["提醒"] = expiring["距离到期天数_最近批次"].apply(
            lambda d: "已临期/180天内到期，建议优先动销或促销" if d <= 180 else "1年内到期，建议关注销售节奏"
        )
        _write_df(
            ws_exp,
            expiring[["货品名称", "货品编号", "汇总", "批次号", "过期日期", "距离到期天数_最近批次", "提醒"]],
            number_format={"汇总": "#,##0"},
        )
    else:
        ws_exp.cell(row=1, column=1, value="无1年内到期的批次记录")

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_PATH)


if __name__ == "__main__":
    tables = clean()
    build_workbook(tables)
    print("仓库列:", tables["warehouses"])
    print(tables["goods"]["品牌"].value_counts())
    print("--- 包装物料按品牌 ---")
    print(tables["packaging"]["品牌"].value_counts())
    print(f"已写入 {OUT_PATH}")
