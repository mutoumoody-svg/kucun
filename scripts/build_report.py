"""
读取 analyze.py 产出的月度SKU表和品类汇总表，生成带图表的Excel周转率分析报告。

固定输出7个sheet（对应交接说明第6节）：
01_月度周转趋势 / 02_库存量趋势 / 03_SKU月度明细 /
04_积压排行 / 05_补货预警 / 06_状态分布 / 07_N周变化

样式：深色表头(#1F2937)，状态颜色编码（快速=绿 正常=蓝 偏慢=橙 积压=红），
千分位数字格式，表头冻结。
"""
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

DATA_DIR = Path(__file__).resolve().parent.parent / "output"
SNAPSHOTS_PATH = DATA_DIR / "snapshots_long.csv"
SKU_MONTHLY_PATH = DATA_DIR / "data_sku_monthly_turnover.csv"
CATEGORY_SUMMARY_PATH = DATA_DIR / "data_category_monthly_summary.csv"
REPORT_PATH = DATA_DIR / "库存周转率分析报告.xlsx"

HEADER_FILL = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)

STATUS_COLORS = {
    "快速": "C6EFCE",  # 绿
    "正常": "BDD7EE",  # 蓝
    "偏慢": "FCE4D6",  # 橙
    "积压": "FFC7CE",  # 红
    "未知": "D9D9D9",
}


def _write_header(ws, headers, row=1):
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=col, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center")
    ws.freeze_panes = ws.cell(row=row + 1, column=1)


def _autosize(ws, n_cols):
    for col in range(1, n_cols + 1):
        letter = get_column_letter(col)
        max_len = max(
            (len(str(ws.cell(row=r, column=col).value or "")) for r in range(1, ws.max_row + 1)),
            default=10,
        )
        ws.column_dimensions[letter].width = min(max(max_len + 2, 10), 40)


def _write_df(ws, df: pd.DataFrame, start_row=1, number_format=None):
    _write_header(ws, list(df.columns), row=start_row)
    for i, (_, r) in enumerate(df.iterrows(), start=start_row + 1):
        for j, col in enumerate(df.columns, start=1):
            cell = ws.cell(row=i, column=j, value=r[col])
            if number_format and col in number_format:
                cell.number_format = number_format[col]
    _autosize(ws, len(df.columns))


def sheet_01_monthly_trend(wb, category_summary: pd.DataFrame):
    ws = wb.create_sheet("01_月度周转趋势")
    pivot = category_summary.pivot(index="month", columns="category", values="median_turnover_months")
    pivot = pivot.reset_index()
    _write_df(ws, pivot, number_format={c: "0.00" for c in pivot.columns if c != "month"})

    chart = LineChart()
    chart.title = "各品类中位周转月数趋势"
    n_rows, n_cols = pivot.shape
    data = Reference(ws, min_col=2, max_col=n_cols, min_row=1, max_row=n_rows + 1)
    cats = Reference(ws, min_col=1, min_row=2, max_row=n_rows + 1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, f"A{n_rows + 4}")

    start_row = n_rows + 25
    status_pivot = category_summary[
        ["category", "month", "count_快速", "count_正常", "count_偏慢", "count_积压"]
    ]
    _write_df(ws, status_pivot, start_row=start_row)


def sheet_02_qty_trend(wb, category_summary: pd.DataFrame):
    ws = wb.create_sheet("02_库存量趋势")
    pivot = category_summary.pivot(index="month", columns="category", values="total_qty").reset_index()
    _write_df(ws, pivot, number_format={c: "#,##0" for c in pivot.columns if c != "month"})

    chart = BarChart()
    chart.type = "col"
    chart.grouping = "clustered"
    chart.title = "品类月均库存总量"
    n_rows, n_cols = pivot.shape
    data = Reference(ws, min_col=2, max_col=n_cols, min_row=1, max_row=n_rows + 1)
    cats = Reference(ws, min_col=1, min_row=2, max_row=n_rows + 1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, f"A{n_rows + 4}")


def sheet_03_sku_detail(wb, monthly_sku: pd.DataFrame):
    ws = wb.create_sheet("03_SKU月度明细")
    latest_month = monthly_sku["month"].max()
    latest = monthly_sku[monthly_sku["month"] == latest_month].sort_values(
        "turnover_months", ascending=False
    )
    order = latest.sort_values(["category", "turnover_months"], ascending=[True, False])[
        ["sku", "category"]
    ].drop_duplicates()

    df = monthly_sku.merge(order, on=["sku", "category"], how="inner")
    pivot_qty = df.pivot_table(index=["sku", "name", "category"], columns="month", values="avg_qty").reset_index()
    _write_df(ws, pivot_qty, number_format={c: "#,##0" for c in pivot_qty.columns if c not in ("sku", "name", "category")})


def sheet_04_backlog_ranking(wb, monthly_sku: pd.DataFrame):
    ws = wb.create_sheet("04_积压排行")
    latest_month = monthly_sku["month"].max()
    df = monthly_sku[(monthly_sku["month"] == latest_month) & (monthly_sku["status"] == "积压")]
    df = df.sort_values("turnover_months", ascending=False).copy()
    df["建议"] = "周转超过6个月，建议清库/促销/下架"
    cols = ["sku", "name", "category", "avg_qty", "turnover_months", "建议"]
    _write_df(ws, df[cols], number_format={"turnover_months": "0.0", "avg_qty": "#,##0"})

    for i, row in enumerate(df.itertuples(), start=2):
        status_col = cols.index("turnover_months") + 1
        ws.cell(row=i, column=status_col).fill = PatternFill(
            start_color=STATUS_COLORS["积压"], end_color=STATUS_COLORS["积压"], fill_type="solid"
        )

    if len(df):
        chart = BarChart()
        chart.type = "bar"
        chart.title = "积压SKU周转月数排行"
        n_rows = len(df)
        data = Reference(ws, min_col=5, min_row=1, max_row=n_rows + 1)
        cats = Reference(ws, min_col=1, min_row=2, max_row=n_rows + 1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        ws.add_chart(chart, f"H2")


def sheet_05_restock_alert(wb, monthly_sku: pd.DataFrame):
    ws = wb.create_sheet("05_补货预警")
    latest_month = monthly_sku["month"].max()
    df = monthly_sku[(monthly_sku["month"] == latest_month) & (monthly_sku["turnover_months"] <= 3)]
    df = df.sort_values("turnover_months").copy()

    def urgency(t):
        if t <= 0.5:
            return "紧急"
        if t <= 1:
            return "高"
        return "中"

    df["紧迫度"] = df["turnover_months"].apply(urgency)
    df["建议"] = "周转月数偏低，关注断货风险，建议尽快补货"
    cols = ["sku", "name", "category", "avg_qty", "turnover_months", "紧迫度", "建议"]
    _write_df(ws, df[cols], number_format={"turnover_months": "0.0", "avg_qty": "#,##0"})

    if len(df):
        chart = BarChart()
        chart.type = "bar"
        chart.title = "补货预警SKU排行"
        n_rows = len(df)
        data = Reference(ws, min_col=5, min_row=1, max_row=n_rows + 1)
        cats = Reference(ws, min_col=1, min_row=2, max_row=n_rows + 1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        ws.add_chart(chart, "H2")


def sheet_06_status_distribution(wb, category_summary: pd.DataFrame):
    ws = wb.create_sheet("06_状态分布")
    latest_month = category_summary["month"].max()
    latest = category_summary[category_summary["month"] == latest_month]
    cols = ["category", "count_快速", "count_正常", "count_偏慢", "count_积压"]
    _write_df(ws, latest[cols])

    for i, cat in enumerate(latest["category"], start=2):
        chart = PieChart()
        chart.title = f"{cat} 状态分布"
        n_row = i
        data = Reference(ws, min_col=2, max_col=5, min_row=1, max_row=n_row)
        cats = Reference(ws, min_col=2, max_col=5, min_row=1, max_row=1)
        chart.add_data(data, titles_from_data=False)
        anchor_col = "H" if i % 2 == 0 else "P"
        ws.add_chart(chart, f"{anchor_col}{2 + (i % 2) * 12}")


def sheet_07_weekly_change(wb, snapshots_long: pd.DataFrame):
    ws = wb.create_sheet("07_周变化")
    df = snapshots_long.copy()
    pivot = df.pivot_table(index=["sku", "name", "category"], columns="date", values="qty").reset_index()
    _write_df(ws, pivot, number_format={c: "#,##0" for c in pivot.columns if c not in ("sku", "name", "category")})


def build_report():
    monthly_sku = pd.read_csv(SKU_MONTHLY_PATH)
    category_summary = pd.read_csv(CATEGORY_SUMMARY_PATH)

    wb = Workbook()
    wb.remove(wb.active)

    sheet_01_monthly_trend(wb, category_summary)
    sheet_02_qty_trend(wb, category_summary)
    sheet_03_sku_detail(wb, monthly_sku)
    sheet_04_backlog_ranking(wb, monthly_sku)
    sheet_05_restock_alert(wb, monthly_sku)
    sheet_06_status_distribution(wb, category_summary)

    if SNAPSHOTS_PATH.exists():
        snapshots_long = pd.read_csv(SNAPSHOTS_PATH)
        sheet_07_weekly_change(wb, snapshots_long)

    DATA_DIR.mkdir(parents=True, exist_ok=True)
    wb.save(REPORT_PATH)
    print(f"报告已生成：{REPORT_PATH}")


if __name__ == "__main__":
    build_report()
