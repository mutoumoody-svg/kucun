"""
统一的"最新库存整理"脚本——只看最新一期快照，输出一个固定文件名的Excel
（output/库存整理.xlsx），每次运行直接覆盖上一次的结果，不再按日期分别留存
260624库存整理.xlsx / 0629库存整理.xlsx这种多个文件。

库存快照来源有两种不同格式，登记在SNAPSHOT_REGISTRY里（按时间顺序）：
- "260624_style"：单Sheet1，按仓库逐行展开，列名 货品编号/货品名称/仓库/库存量/
  可发库存?，例如260624原始数据.xlsx。这种格式本身不含保质期字段。
- "erp_style"：ERP批次导出，可能拆多个sheet（蜂蜜/sttoke/库存6.1等），列名
  商家编码/货品名称/仓库/正常库存/批次号/生产日期/过期日期/距离到期天数/次品标记，
  例如副本库存6.29.xlsx。这种格式本身就带保质期字段。

以后每收到一期新快照，只要在SNAPSHOT_REGISTRY末尾加一行（日期、格式、文件路径），
脚本会自动挑日期最新的一条来跑，不需要改其他逻辑。

整理逻辑（跟之前两个独立脚本一致）：
1. 同一SKU不同仓库的库存量横向展开成并列的列，最后加一列"汇总"
2. 按品牌分类（STTOKE/慕咖/食品/周边商品），商品和包装物料分开列，各自再按品牌分sheet
3. 食品品牌附保质期信息（260624_style格式没有原生保质期字段时，从最近的ERP批次
   导出里按SKU join过来）
4. 食品/慕咖品牌按人工指定的固定顺序展示，慕咖品牌最上面单独置顶几项挂耳咖啡
5. 出库数量是"当月累计出库"（月初到本期快照的库存量差），不是跟上一期快照比较
6. 货品名称放第一列，同规格关键词（12OZ/16OZ等）的商品排在一起
7. "库存状态"单元格按状态上色（蓝色系，积压用粉红色突出），"汇总"列文字加粗放大
"""
import json
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from categorize import (
    BRAND_ORDER,
    MOODY_TOP_ITEMS,
    MOODY_TOP_ITEMS_SKU,
    NAME_OVERRIDE,
    SKU_ALIAS,
    STATUS_COLORS,
    classify_brand,
    classify_material,
    extract_group_keyword,
    group_sort_key,
    sku_sort_key,
)
from monthly_outbound import cumulative_month_outbound
from turnover_lookup import load_latest_turnover

RAW_DIR = Path(__file__).resolve().parent.parent / "raw_data"
OUT_PATH = Path(__file__).resolve().parent.parent / "output" / "库存整理.xlsx"
REGISTRY_PATH = RAW_DIR / "snapshot_registry.json"
NAME_REGISTRY_PATH = RAW_DIR / "name_registry.json"


def load_name_registry() -> dict:
    """加载持久化的 SKU->名称 注册表"""
    if not NAME_REGISTRY_PATH.exists():
        return {}
    with open(NAME_REGISTRY_PATH, "r", encoding="utf-8") as f:
        return json.load(f)


def save_name_registry(registry: dict) -> None:
    with open(NAME_REGISTRY_PATH, "w", encoding="utf-8") as f:
        json.dump(registry, f, ensure_ascii=False, indent=2)


def update_name_registry(sku_name_pairs: dict) -> None:
    """用新读到的 SKU->名称 更新注册表（新名称覆盖旧名称）"""
    registry = load_name_registry()
    registry.update({k: v for k, v in sku_name_pairs.items() if v and str(v).strip()})
    save_name_registry(registry)

# 已知保质期来源（用于补全260624_style快照没有的保质期字段），按时间顺序，
# 同一SKU在多个来源里都出现时优先用较新来源
EXPIRY_FALLBACK_SOURCES = [
    (RAW_DIR / "副本库存6.1(1).xlsx", "库存6.1"),
    (RAW_DIR / "副本库存6.29.xlsx", "蜂蜜"),
]

# 已知库存快照登记表的初始种子数据（第一次运行时如果 raw_data/snapshot_registry.json
# 还不存在，会用这份种子数据初始化它）。登记表格式：(日期, 格式, 文件名, [erp格式下
# 要读的sheet列表])。文件名是相对raw_data目录的文件名，不是绝对路径，方便挪动项目。
SEED_REGISTRY = [
    ("2026-06-24", "260624_style", "260624原始数据.xlsx", None),
    ("2026-06-29", "erp_style", "副本库存6.29.xlsx", ["蜂蜜", "sttoke"]),
    ("2026-06-30", "erp_style", "副本库存6.30.xlsx", ["蜂蜜", "慕咖sttoke"]),
]


def load_registry() -> list[tuple]:
    """从JSON文件读取快照登记表，文件不存在时用SEED_REGISTRY初始化"""
    if not REGISTRY_PATH.exists():
        save_registry(SEED_REGISTRY)
        return list(SEED_REGISTRY)
    with open(REGISTRY_PATH, "r", encoding="utf-8") as f:
        data = json.load(f)
    return [(e["date"], e["format"], e["filename"], e["sheets"]) for e in data]


def save_registry(entries: list[tuple]) -> None:
    data = [
        {"date": d, "format": fmt, "filename": fn, "sheets": sheets}
        for d, fmt, fn, sheets in entries
    ]
    REGISTRY_PATH.parent.mkdir(parents=True, exist_ok=True)
    with open(REGISTRY_PATH, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def detect_format_and_sheets(path: Path) -> tuple[str, list[str] | None]:
    """嗅探一个新上传的xlsx是260624_style还是erp_style，erp_style还要找出
    哪些sheet是真正的库存数据sheet（含 商家编码/正常库存 列）"""
    xls = pd.ExcelFile(path)
    if "Sheet1" in xls.sheet_names:
        cols = pd.read_excel(path, sheet_name="Sheet1", nrows=1).columns
        if "货品编号" in cols and "库存量" in cols:
            return "260624_style", None

    erp_sheets = []
    for sheet in xls.sheet_names:
        cols = pd.read_excel(path, sheet_name=sheet, nrows=1).columns
        if "商家编码" in cols and "正常库存" in cols:
            erp_sheets.append(sheet)
    if erp_sheets:
        return "erp_style", erp_sheets

    raise ValueError(f"无法识别文件格式: {path.name}，既不是260624_style也不是erp_style")


def detect_snapshot_date(path: Path, fmt: str, sheets: list[str] | None, fallback_date: str | None = None) -> str:
    """erp_style可以直接从'库存日期'列读出日期；260624_style没有这个字段，
    只能用上传时提供的日期，没提供就用文件最后修改时间"""
    if fmt == "erp_style":
        df = pd.read_excel(path, sheet_name=sheets[0], nrows=1)
        if "库存日期" in df.columns:
            return str(df["库存日期"].iloc[0])
    if fallback_date:
        return fallback_date
    import datetime

    return datetime.date.fromtimestamp(path.stat().st_mtime).isoformat()


def register_snapshot(path: Path, date_override: str | None = None) -> tuple[str, str, str, list[str] | None]:
    """供网页上传调用：识别格式、提取日期、写入raw_data并登记进registry，
    返回新增的登记条目"""
    fmt, sheets = detect_format_and_sheets(path)
    date = detect_snapshot_date(path, fmt, sheets, fallback_date=date_override)

    entries = load_registry()
    entries = [e for e in entries if e[2] != path.name]  # 同名文件覆盖旧记录
    entry = (date, fmt, path.name, sheets)
    entries.append(entry)
    save_registry(entries)
    return entry

HEADER_FILL = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
TOTAL_FONT = Font(bold=True, size=12, color="000000")


def _write_header(ws, headers, row=1):
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=col, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center")
    ws.freeze_panes = ws.cell(row=row + 1, column=1)


def _autosize(ws, n_cols):
    for col in range(1, n_cols + 1):
        letter = get_column_letter(col)
        max_len = max(
            (len(str(ws.cell(row=r, column=col).value or "")) for r in range(1, ws.max_row + 1)),
            default=10,
        )
        ws.column_dimensions[letter].width = min(max(max_len + 2, 10), 50)


def _write_rows(ws, df: pd.DataFrame, columns: list[str], start_row: int, number_format: dict | None = None) -> int:
    status_col_idx = columns.index("库存状态") + 1 if "库存状态" in columns else None
    total_col_idx = columns.index("汇总") + 1 if "汇总" in columns else None
    row = start_row
    for _, r in df.iterrows():
        for j, col in enumerate(columns, start=1):
            cell = ws.cell(row=row, column=j, value=r[col])
            if number_format and col in number_format:
                cell.number_format = number_format[col]
            if status_col_idx and j == status_col_idx:
                color = STATUS_COLORS.get(r["库存状态"])
                if color:
                    cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            if total_col_idx and j == total_col_idx:
                cell.font = TOTAL_FONT
        row += 1
    return row


def _write_df(ws, df: pd.DataFrame, number_format: dict | None = None):
    columns = list(df.columns)
    _write_header(ws, columns)
    _write_rows(ws, df, columns, start_row=2, number_format=number_format)
    _autosize(ws, len(columns))


def _write_df_with_gap(ws, df_top: pd.DataFrame, df_rest: pd.DataFrame, gap_rows: int, number_format: dict | None = None):
    columns = list(df_top.columns)
    _write_header(ws, columns)
    next_row = _write_rows(ws, df_top, columns, start_row=2, number_format=number_format)
    _write_rows(ws, df_rest, columns, start_row=next_row + gap_rows, number_format=number_format)
    _autosize(ws, len(columns))


def pick_latest_snapshot():
    entries = load_registry()
    date, fmt, filename, sheets = max(entries, key=lambda x: x[0])
    return date, fmt, RAW_DIR / filename, sheets


def load_260624_style(path: Path) -> pd.DataFrame:
    """返回统一长表：sku, name, warehouse, qty（不含原生保质期）"""
    df = pd.read_excel(path, sheet_name="Sheet1")
    df["货品编号"] = df["货品编号"].replace(SKU_ALIAS)
    result = df.rename(
        columns={"货品编号": "sku", "货品名称": "name", "仓库": "warehouse", "库存量": "qty"}
    )
    if "name" not in result.columns:
        result["name"] = ""

    # 260624格式只读注册表补全名称，不写入（名称权威来源只有ERP文件）
    name_reg = load_name_registry()
    result["name"] = result.apply(
        lambda r: name_reg.get(r["sku"], r["name"]), axis=1
    )
    return result[["sku", "name", "warehouse", "qty"]]


def load_erp_style(path: Path, sheets: list[str]) -> tuple[pd.DataFrame, pd.DataFrame]:
    """返回 (统一长表: sku,name,warehouse,qty,  保质期表: sku,批次号,生产日期,过期日期,距离到期天数_最近批次)"""
    frames = []
    for sheet in sheets:
        d = pd.read_excel(path, sheet_name=sheet)
        if "次品标记" in d.columns:
            d = d[d["次品标记"] != 1]
        d["商家编码"] = d["商家编码"].replace(SKU_ALIAS)
        frames.append(d)
    raw = pd.concat(frames, ignore_index=True)

    long_df = raw.rename(
        columns={"商家编码": "sku", "货品名称": "name", "仓库": "warehouse", "正常库存": "qty"}
    )[["sku", "name", "warehouse", "qty"]]

    # 每次读ERP文件都把最新的SKU->名称写入注册表
    name_pairs = dict(zip(long_df["sku"], long_df["name"]))
    update_name_registry(name_pairs)

    has_real_expiry = raw["生产日期"].astype(str) != "0000-00-00 00:00:00"
    expiry_rows = raw[has_real_expiry].sort_values(["商家编码", "距离到期天数"])
    expiry = expiry_rows.groupby("商家编码").agg(
        批次号=("批次号", lambda s: " / ".join(str(x) for x in s)),
        生产日期=("生产日期", "first"),
        过期日期=("过期日期", "first"),
        距离到期天数_最近批次=("距离到期天数", "min"),
    ).reset_index().rename(columns={"商家编码": "sku"})

    return long_df, expiry


def load_expiry_fallback() -> pd.DataFrame:
    """260624_style快照没有原生保质期字段时，从最近的ERP批次导出里按SKU补全保质期"""
    frames = []
    for path, sheet in EXPIRY_FALLBACK_SOURCES:
        if not path.exists():
            continue
        erp = pd.read_excel(path, sheet_name=sheet)
        snapshot_date = str(erp["库存日期"].iloc[0])
        real = erp[erp["生产日期"].astype(str) != "0000-00-00 00:00:00"].copy()
        real["sku"] = real["商家编码"].replace(SKU_ALIAS)
        real["保质期数据日期"] = snapshot_date
        real["来源顺序"] = len(frames)
        frames.append(real)

    if not frames:
        return pd.DataFrame(columns=["sku", "批次号", "生产日期", "过期日期", "距离到期天数_最近批次", "保质期数据日期"])

    combined = pd.concat(frames, ignore_index=True)
    combined = combined.sort_values("来源顺序").drop_duplicates(subset=["sku", "批次号"], keep="last")
    combined = combined.sort_values(["sku", "距离到期天数"])
    return combined.groupby("sku").agg(
        批次号=("批次号", lambda s: " / ".join(str(x) for x in s)),
        生产日期=("生产日期", "first"),
        过期日期=("过期日期", "first"),
        距离到期天数_最近批次=("距离到期天数", "min"),
        保质期数据日期=("保质期数据日期", "last"),
    ).reset_index()


def _ensure_name_registry() -> None:
    """首次运行或注册表为空时，从所有历史ERP文件里批量提取名称写入注册表。"""
    if NAME_REGISTRY_PATH.exists() and load_name_registry():
        return
    registry = load_registry()
    for date, fmt, filename, sheets in registry:
        if fmt != "erp_style" or not sheets:
            continue
        path = RAW_DIR / filename
        if not path.exists():
            continue
        try:
            frames = []
            for sheet in sheets:
                d = pd.read_excel(path, sheet_name=sheet)
                if "次品标记" in d.columns:
                    d = d[d["次品标记"] != 1]
                d["商家编码"] = d["商家编码"].replace(SKU_ALIAS)
                frames.append(d)
            raw = pd.concat(frames, ignore_index=True)
            long_df = raw.rename(columns={"商家编码": "sku", "货品名称": "name"})[["sku", "name"]]
            pairs = {r["sku"]: r["name"] for _, r in long_df.iterrows() if r["name"]}
            update_name_registry(pairs)
        except Exception:
            pass


def clean() -> dict:
    _ensure_name_registry()
    snapshot_date, fmt, path, sheets = pick_latest_snapshot()

    if fmt == "260624_style":
        long_df = load_260624_style(path)
        expiry = load_expiry_fallback()
    else:
        long_df, expiry = load_erp_style(path, sheets)
        expiry["保质期数据日期"] = snapshot_date

    warehouses = sorted(long_df["warehouse"].unique())

    wide = long_df.pivot_table(
        index=["sku", "name"], columns="warehouse", values="qty", aggfunc="sum", fill_value=0
    ).reset_index()
    wide["汇总"] = wide[warehouses].sum(axis=1)
    wide = wide.rename(columns={"sku": "货品编号", "name": "货品名称"})
    # 优先用ERP注册表里的最新名称（上传的ERP文件名字最权威），
    # 注册表里没有的SKU才用NAME_OVERRIDE补全，再没有就用文件原始名称
    name_reg = load_name_registry()
    wide["货品名称"] = wide.apply(
        lambda r: name_reg.get(r["货品编号"],
                               NAME_OVERRIDE.get(r["货品编号"], r["货品名称"])), axis=1
    )

    wide["类型"] = wide["货品名称"].apply(classify_material)
    wide["品牌"] = wide["货品名称"].apply(classify_brand)
    wide["规格分组"] = wide["货品名称"].apply(extract_group_keyword)

    wide = wide.merge(expiry, left_on="货品编号", right_on="sku", how="left")
    if "sku" in wide.columns:
        wide = wide.drop(columns=["sku"])

    turnover = load_latest_turnover()
    wide = wide.merge(turnover, left_on="货品编号", right_on="sku", how="left").drop(columns=["sku"])

    outbound = cumulative_month_outbound(snapshot_date)
    wide = wide.merge(outbound, left_on="货品编号", right_on="sku", how="left").drop(columns=["sku"])

    def _sort_weight(r) -> int:
        # 统一按各页面SKU顺序表排序；不在表里的新SKU排到末尾
        return sku_sort_key(r["货品编号"], r["类型"], r["品牌"])

    def reorder(d: pd.DataFrame) -> pd.DataFrame:
        d = d.copy()
        d["_排序权重"] = d.apply(_sort_weight, axis=1)
        d = d.sort_values(
            ["_排序权重", "距离到期天数_最近批次", "汇总"],
            ascending=[True, True, False],
            na_position="last",
        ).drop(columns=["_排序权重"])
        cols = ["货品名称", "规格分组"] + [c for c in d.columns if c not in ("货品名称", "规格分组")]
        return d[cols]

    goods = reorder(wide[wide["类型"] == "商品"])
    packaging = reorder(wide[wide["类型"] == "包装物料"])

    low_stock = pd.DataFrame()
    if fmt == "260624_style":
        low_stock = goods[(goods["汇总"] > 0) & (goods["汇总"] <= 10)]

    expiring_soon = goods[
        (goods["品牌"] == "食品") & goods["距离到期天数_最近批次"].notna() & (goods["距离到期天数_最近批次"] <= 365)
    ]

    return {
        "goods": goods,
        "packaging": packaging,
        "low_stock": low_stock,
        "expiring_soon": expiring_soon,
        "warehouses": warehouses,
        "snapshot_date": snapshot_date,
        "format": fmt,
    }


def _brand_cols(brand: str, warehouses: list[str]) -> list[str]:
    base = ["货品名称", "规格分组", "货品编号"] + warehouses + ["汇总"]
    if brand == "食品":
        base += ["批次号", "生产日期", "过期日期", "距离到期天数_最近批次", "保质期数据日期"]
    base += ["累计出库数量", "出库数据周期", "最新月周转月数", "库存状态", "周转数据月份"]
    return base


def build_workbook(tables: dict):
    wb = Workbook()
    wb.remove(wb.active)
    warehouses = tables["warehouses"]
    qty_fmt = {w: "#,##0" for w in warehouses}
    qty_fmt.update({"汇总": "#,##0", "累计出库数量": "#,##0", "最新月周转月数": "0.0"})
    idx = 1

    for brand in BRAND_ORDER:
        sub = tables["goods"][tables["goods"]["品牌"] == brand]
        if sub.empty:
            continue
        ws = wb.create_sheet(f"{idx:02d}_商品-{brand}")
        cols = _brand_cols(brand, warehouses)
        if brand == "慕咖":
            top = sub[sub["货品编号"].isin(MOODY_TOP_ITEMS_SKU)]
            rest = sub[~sub["货品编号"].isin(MOODY_TOP_ITEMS_SKU)]
            _write_df_with_gap(ws, top[cols], rest[cols], gap_rows=2, number_format=qty_fmt)
        else:
            _write_df(ws, sub[cols], number_format=qty_fmt)
        idx += 1

    for brand in BRAND_ORDER:
        sub = tables["packaging"][tables["packaging"]["品牌"] == brand]
        if sub.empty:
            continue
        ws = wb.create_sheet(f"{idx:02d}_包装物料-{brand}")
        cols = ["货品名称", "规格分组", "货品编号"] + warehouses + ["汇总"]
        _write_df(ws, sub[cols], number_format=qty_fmt)
        idx += 1

    if len(tables["low_stock"]):
        ws_low = wb.create_sheet(f"{idx:02d}_低库存预警")
        idx += 1
        low = tables["low_stock"].copy()
        low["提醒"] = "总库存≤10，建议核实是否需要补货/下架"
        _write_df(
            ws_low,
            low[["货品名称", "货品编号", "品牌"] + warehouses + ["汇总", "提醒"]],
            number_format=qty_fmt,
        )

    ws_exp = wb.create_sheet(f"{idx:02d}_临期预警")
    expiring = tables["expiring_soon"].copy()
    if len(expiring):
        expiring["提醒"] = expiring["距离到期天数_最近批次"].apply(
            lambda d: "已临期/180天内到期，建议优先动销或促销" if d <= 180 else "1年内到期，建议关注销售节奏"
        )
        _write_df(
            ws_exp,
            expiring[["货品名称", "货品编号", "汇总", "批次号", "过期日期", "距离到期天数_最近批次", "提醒"]],
            number_format={"汇总": "#,##0"},
        )
    else:
        ws_exp.cell(row=1, column=1, value="无1年内到期的批次记录")

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_PATH)


def run_pipeline() -> dict:
    """跑一遍完整流程（取最新快照->整理->写出Excel），供脚本和网页后端共用。
    返回tables字典，里面附带snapshot_date/format等信息，方便调用方展示状态。"""
    tables = clean()
    build_workbook(tables)
    return tables


if __name__ == "__main__":
    tables = run_pipeline()
    print(f"最新快照日期: {tables['snapshot_date']} (格式: {tables['format']})")
    print("仓库列:", tables["warehouses"])
    print(tables["goods"]["品牌"].value_counts())
    print("--- 包装物料按品牌 ---")
    print(tables["packaging"]["品牌"].value_counts())
    print(f"已写入 {OUT_PATH}")
