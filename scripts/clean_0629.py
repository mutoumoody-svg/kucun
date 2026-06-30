"""
整理 副本库存6.29.xlsx（ERP批次导出，6/29，sheet名为 蜂蜜/sttoke）。

跟260624原始数据.xlsx不同，这个文件本身就是ERP批次格式，每行已经带有
批次号/生产日期/过期日期/距离到期天数/仓库，不需要再去外部文件join保质期。
sttoke sheet里混合了STTOKE和慕咖两个品类的商品（货品名称里能区分），
蜂蜜 sheet都是食品。

整理逻辑：
1. 过滤次品标记==1的记录
2. 同一SKU不同仓库的正常库存横向展开成并列的列（一行一个产品），后面加一列
   "汇总"，不再按仓库拆成多行
3. 按品牌分类（复用 categorize.py）：STTOKE/慕咖/食品/周边商品，不区分商品种类
   —— 杯子、蜂蜜、苹果醋、咖啡只要是同一品牌就放在同一张表里，不按食品/非食品拆开
4. 包装物料同样按品牌单独分表，不与对应品牌的商品混在一张表里
5. 食品品牌的商品附批次号/过期日期/距离到期天数
6. 出库数量是"当月累计出库"（月初到本期快照的累计下降量，中途补货不会冲淡这个数），
   不是单纯跟上一期快照的差值
7. 货品名称放在第一列，同规格关键词（如12OZ/16OZ）的产品排在一起
"""
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from categorize import (
    BRAND_ORDER,
    MOODY_TOP_ITEMS,
    SKU_ALIAS,
    STATUS_COLORS,
    classify_brand,
    classify_material,
    extract_group_keyword,
    food_sort_key,
    group_sort_key,
    moody_sort_key,
)
from monthly_outbound import cumulative_month_outbound
from turnover_lookup import load_latest_turnover

RAW_DIR = Path(__file__).resolve().parent.parent / "raw_data"
SRC_PATH = RAW_DIR / "副本库存6.29.xlsx"
OUT_PATH = Path(__file__).resolve().parent.parent / "output" / "0629库存整理.xlsx"

HEADER_FILL = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def _write_header(ws, headers, row=1):
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=col, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center")
    ws.freeze_panes = ws.cell(row=row + 1, column=1)


def _autosize(ws, n_cols):
    for col in range(1, n_cols + 1):
        letter = get_column_letter(col)
        max_len = max(
            (len(str(ws.cell(row=r, column=col).value or "")) for r in range(1, ws.max_row + 1)),
            default=10,
        )
        ws.column_dimensions[letter].width = min(max(max_len + 2, 10), 50)


TOTAL_FONT = Font(bold=True, size=12, color="000000")


def _write_rows(ws, df: pd.DataFrame, columns: list[str], start_row: int, number_format: dict | None = None) -> int:
    """从start_row开始写数据行（不写表头），返回写完之后的下一行行号"""
    status_col_idx = columns.index("库存状态") + 1 if "库存状态" in columns else None
    total_col_idx = columns.index("汇总") + 1 if "汇总" in columns else None
    row = start_row
    for _, r in df.iterrows():
        for j, col in enumerate(columns, start=1):
            cell = ws.cell(row=row, column=j, value=r[col])
            if number_format and col in number_format:
                cell.number_format = number_format[col]
            if status_col_idx and j == status_col_idx:
                color = STATUS_COLORS.get(r["库存状态"])
                if color:
                    cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            if total_col_idx and j == total_col_idx:
                cell.font = TOTAL_FONT
        row += 1
    return row


def _write_df(ws, df: pd.DataFrame, number_format: dict | None = None):
    columns = list(df.columns)
    _write_header(ws, columns)
    _write_rows(ws, df, columns, start_row=2, number_format=number_format)
    _autosize(ws, len(columns))


def _write_df_with_gap(ws, df_top: pd.DataFrame, df_rest: pd.DataFrame, gap_rows: int, number_format: dict | None = None):
    """先写df_top，空gap_rows行，再接着写df_rest，表头只写一次"""
    columns = list(df_top.columns)
    _write_header(ws, columns)
    next_row = _write_rows(ws, df_top, columns, start_row=2, number_format=number_format)
    _write_rows(ws, df_rest, columns, start_row=next_row + gap_rows, number_format=number_format)
    _autosize(ws, len(columns))


def load_raw() -> pd.DataFrame:
    frames = []
    xls = pd.ExcelFile(SRC_PATH)
    for sheet in xls.sheet_names:
        df = pd.read_excel(SRC_PATH, sheet_name=sheet)
        if "次品标记" in df.columns:
            df = df[df["次品标记"] != 1]
        frames.append(df)
    raw = pd.concat(frames, ignore_index=True)
    raw["商家编码"] = raw["商家编码"].replace(SKU_ALIAS)
    return raw


def clean() -> dict:
    raw = load_raw()
    snapshot_date = str(raw["库存日期"].iloc[0])
    has_real_expiry = raw["生产日期"].astype(str) != "0000-00-00 00:00:00"
    warehouses = sorted(raw["仓库"].unique())

    wide = raw.pivot_table(
        index=["商家编码", "货品名称"], columns="仓库", values="正常库存", aggfunc="sum", fill_value=0
    ).reset_index()
    wide["汇总"] = wide[warehouses].sum(axis=1)

    batch_count = raw.groupby("商家编码")["批次号"].nunique().rename("涉及批次数")
    wide = wide.merge(batch_count, on="商家编码", how="left")

    wide["类型"] = wide["货品名称"].apply(classify_material)
    wide["品牌"] = wide["货品名称"].apply(classify_brand)
    wide["规格分组"] = wide["货品名称"].apply(extract_group_keyword)

    expiry_rows = raw[has_real_expiry].sort_values(["商家编码", "距离到期天数"])
    expiry = expiry_rows.groupby("商家编码").agg(
        批次号=("批次号", lambda s: " / ".join(str(x) for x in s)),
        生产日期=("生产日期", "first"),
        过期日期=("过期日期", "first"),
        距离到期天数_最近批次=("距离到期天数", "min"),
    ).reset_index()
    wide = wide.merge(expiry, on="商家编码", how="left")
    wide["保质期数据日期"] = snapshot_date

    turnover = load_latest_turnover()
    wide = wide.merge(turnover, left_on="商家编码", right_on="sku", how="left").drop(columns=["sku"])

    outbound = cumulative_month_outbound(snapshot_date)
    wide = wide.merge(outbound, left_on="商家编码", right_on="sku", how="left").drop(columns=["sku"])

    def _sort_weight(r) -> int:
        if r["品牌"] == "食品":
            return food_sort_key(r["货品名称"])
        if r["品牌"] == "慕咖":
            if r["货品名称"] in MOODY_TOP_ITEMS:
                return MOODY_TOP_ITEMS.index(r["货品名称"])
            return moody_sort_key(r["货品名称"])
        return group_sort_key(r["规格分组"])

    def reorder(d: pd.DataFrame) -> pd.DataFrame:
        d = d.copy()
        # 食品/慕咖品牌按人工指定的固定顺序排，其余品牌按规格分组排
        d["_排序权重"] = d.apply(_sort_weight, axis=1)
        d = d.sort_values(
            ["_排序权重", "规格分组", "距离到期天数_最近批次", "汇总"],
            ascending=[True, True, True, False],
            na_position="last",
        ).drop(columns=["_排序权重"])
        cols = ["货品名称", "规格分组"] + [c for c in d.columns if c not in ("货品名称", "规格分组")]
        return d[cols]

    goods = reorder(wide[wide["类型"] == "商品"])
    packaging = reorder(wide[wide["类型"] == "包装物料"])

    expiring_soon = goods[
        (goods["品牌"] == "食品") & goods["距离到期天数_最近批次"].notna() & (goods["距离到期天数_最近批次"] <= 365)
    ]

    return {
        "goods": goods,
        "packaging": packaging,
        "expiring_soon": expiring_soon,
        "snapshot_date": snapshot_date,
        "warehouses": warehouses,
    }


def _brand_cols(brand: str, warehouses: list[str]) -> list[str]:
    base = ["货品名称", "规格分组", "商家编码"] + warehouses + ["汇总", "涉及批次数"]
    if brand == "食品":
        base += ["批次号", "生产日期", "过期日期", "距离到期天数_最近批次", "保质期数据日期"]
    base += ["累计出库数量", "出库数据周期", "最新月周转月数", "库存状态", "周转数据月份"]
    return base


def build_workbook(tables: dict):
    wb = Workbook()
    wb.remove(wb.active)
    warehouses = tables["warehouses"]
    qty_fmt = {w: "#,##0" for w in warehouses}
    qty_fmt.update({"汇总": "#,##0", "累计出库数量": "#,##0", "最新月周转月数": "0.0"})
    idx = 1

    for brand in BRAND_ORDER:
        sub = tables["goods"][tables["goods"]["品牌"] == brand]
        if sub.empty:
            continue
        ws = wb.create_sheet(f"{idx:02d}_商品-{brand}")
        cols = _brand_cols(brand, warehouses)
        if brand == "慕咖":
            top = sub[sub["货品名称"].isin(MOODY_TOP_ITEMS)]
            rest = sub[~sub["货品名称"].isin(MOODY_TOP_ITEMS)]
            _write_df_with_gap(ws, top[cols], rest[cols], gap_rows=2, number_format=qty_fmt)
        else:
            _write_df(ws, sub[cols], number_format=qty_fmt)
        idx += 1

    for brand in BRAND_ORDER:
        sub = tables["packaging"][tables["packaging"]["品牌"] == brand]
        if sub.empty:
            continue
        ws = wb.create_sheet(f"{idx:02d}_包装物料-{brand}")
        cols = ["货品名称", "规格分组", "商家编码"] + warehouses + ["汇总", "涉及批次数"]
        _write_df(ws, sub[cols], number_format=qty_fmt)
        idx += 1

    ws_exp = wb.create_sheet(f"{idx:02d}_临期预警")
    expiring = tables["expiring_soon"].copy()
    if len(expiring):
        expiring["提醒"] = expiring["距离到期天数_最近批次"].apply(
            lambda d: "已临期/180天内到期，建议优先动销或促销" if d <= 180 else "1年内到期，建议关注销售节奏"
        )
        _write_df(
            ws_exp,
            expiring[["货品名称", "商家编码", "汇总", "批次号", "过期日期", "距离到期天数_最近批次", "提醒"]],
            number_format={"汇总": "#,##0"},
        )
    else:
        ws_exp.cell(row=1, column=1, value="无1年内到期的批次记录")

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_PATH)


if __name__ == "__main__":
    tables = clean()
    build_workbook(tables)
    print(f"快照日期: {tables['snapshot_date']}")
    print("仓库列:", tables["warehouses"])
    print(tables["goods"]["品牌"].value_counts())
    print("--- 包装物料按品牌 ---")
    print(tables["packaging"]["品牌"].value_counts())
    print(f"已写入 {OUT_PATH}")
